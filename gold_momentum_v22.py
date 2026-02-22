#!/usr/bin/env python3
"""Gold Momentum Scoring System v2.2 — JP Trust Learning"""

import pandas as pd
import numpy as np
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import os, sys

# ── CONFIG ──
ROLLING_WINDOW = 252
LOOKBACK = {'1W': 5, '1M': 21, '3M': 63, '6M': 126, '1Y': 252}
WEIGHTS = {'1Y': 0.30, '6M': 0.25, '3M': 0.20, '1M': 0.15, '1W': 0.10}
WEIGHT_ORDER = ['1Y', '6M', '3M', '1M', '1W']

RUN_TS = datetime.now(timezone.utc)
AS_OF = RUN_TS.strftime("%d/%m/%Y %H:%M UTC")
TS_FILE = RUN_TS.strftime("%d%m%Y_%H%M")

# ── LOAD DATA ──
csv_path = os.path.join(os.path.dirname(__file__), 'gold_prices.csv')
df = pd.read_csv(csv_path, encoding='utf-8-sig')
df.columns = ['Date', 'Open', 'High', 'Low', 'Close', 'Volume']
df['Date'] = pd.to_datetime(df['Date'])
df = df.sort_values('Date').reset_index(drop=True)

# ── BASE DATES ──
BD2_idx = len(df) - 1
BD1_idx = len(df) - 6
BD1_date = df.iloc[BD1_idx]['Date']
BD2_date = df.iloc[BD2_idx]['Date']

print(f"Base Date 1: {BD1_date.strftime('%Y-%m-%d')} (idx={BD1_idx})")
print(f"Base Date 2: {BD2_date.strftime('%Y-%m-%d')} (idx={BD2_idx})")
print(f"Total rows: {len(df)}")

# ══════════════════════════════════════════════════════
# SCORING FUNCTIONS
# ══════════════════════════════════════════════════════

def compute_return(closes, end_idx, period_days):
    start_idx = end_idx - period_days
    if start_idx < 0:
        return None
    return (closes[end_idx] - closes[start_idx]) / closes[start_idx] * 100

def rolling_percentile(series_values, current_val, window=ROLLING_WINDOW):
    valid = series_values[~np.isnan(series_values)]
    if len(valid) < 10:
        return 50.0
    count_below = np.sum(valid < current_val)
    return count_below / (len(valid) - 1) * 100 if len(valid) > 1 else 50.0

def calc_return_percentiles(df, base_idx):
    closes = df['Close'].values
    results = {}
    for period, days in LOOKBACK.items():
        current_ret = compute_return(closes, base_idx, days)
        if current_ret is None:
            results[period] = {'return': 0, 'percentile': 50}
            continue
        # Build rolling returns for the same period over last 252 days
        rolling_rets = []
        start = max(0, base_idx - ROLLING_WINDOW)
        for i in range(start, base_idx):
            r = compute_return(closes, i, days)
            if r is not None:
                rolling_rets.append(r)
        pctl = rolling_percentile(np.array(rolling_rets), current_ret) if rolling_rets else 50
        results[period] = {'return': current_ret, 'percentile': pctl}
    return results

def calc_volume_percentiles(df, base_idx):
    volumes = df['Volume'].values
    results = {}
    for period, days in LOOKBACK.items():
        end = base_idx + 1
        start = end - days
        if start < 0:
            results[period] = {'volume': 0, 'percentile': 50}
            continue
        current_vol = np.sum(volumes[start:end])
        # Rolling
        rolling_vols = []
        roll_start = max(0, base_idx - ROLLING_WINDOW)
        for i in range(roll_start, base_idx):
            s = i + 1 - days
            if s < 0:
                continue
            rolling_vols.append(np.sum(volumes[s:i+1]))
        pctl = rolling_percentile(np.array(rolling_vols), current_vol) if rolling_vols else 50
        results[period] = {'volume': current_vol, 'percentile': pctl}
    return results

def weighted_percentile(pctl_dict):
    return sum(pctl_dict[p]['percentile'] * WEIGHTS[p] for p in WEIGHT_ORDER)

def d1_score(wp): return wp / 100 * 20
def d2_score(wp): return wp / 100 * 20

def calc_rsi(df, base_idx, period=14):
    start = base_idx - 29
    if start < 0: start = 0
    closes = df['Close'].values[start:base_idx+1]
    if len(closes) < 2:
        return 50
    deltas = np.diff(closes)
    gains = np.where(deltas > 0, deltas, 0)
    losses = np.where(deltas < 0, -deltas, 0)
    last_n = min(period, len(gains))
    avg_gain = np.mean(gains[-last_n:])
    avg_loss = np.mean(losses[-last_n:])
    if avg_loss == 0:
        return 100.0
    rs = avg_gain / avg_loss
    return 100 - (100 / (1 + rs))

def d3_score(rsi):
    if 50 <= rsi <= 70: return 20
    if 40 <= rsi < 50: return 16
    if 70 < rsi <= 80: return 14
    if 30 <= rsi < 40: return 12
    if rsi > 80: return 10
    return 6  # < 30

def calc_ma(df, base_idx, window):
    start = base_idx + 1 - window
    if start < 0: return None
    return np.mean(df['Close'].values[start:base_idx+1])

def d4_score(price, ma50, ma200):
    pts = 0
    if ma50 is not None and price > ma50: pts += 7
    if ma200 is not None and price > ma200: pts += 7
    if ma50 is not None and ma200 is not None and ma50 > ma200: pts += 6
    return min(pts, 20)

def calc_volatility(df, base_idx):
    start = base_idx - 20
    if start < 0: start = 0
    closes = df['Close'].values[start:base_idx+1]
    if len(closes) < 2:
        return 0
    rets = np.diff(closes) / closes[:-1]
    return np.std(rets) * np.sqrt(252) * 100

def d5_score(vol):
    if vol <= 20: return 20
    if vol <= 30: return 18
    if vol <= 40: return 14
    if vol <= 50: return 11
    if vol <= 60: return 8
    if vol <= 80: return 5
    return 2

def calc_penalties(df, base_idx):
    closes = df['Close'].values
    ret_1y = compute_return(closes, base_idx, 252)
    ret_6m = compute_return(closes, base_idx, 126)
    ret_1m = compute_return(closes, base_idx, 21)
    ret_1w = compute_return(closes, base_idx, 5)
    
    # defaults
    if ret_1y is None: ret_1y = 0
    if ret_6m is None: ret_6m = 0
    if ret_1m is None: ret_1m = 0
    if ret_1w is None: ret_1w = 0
    
    # Reversal
    reversal_pen = 0
    reversal_flag = ""
    strong = (ret_1y > 20 and ret_1m < -5 and ret_1w < -3)
    mild = ((ret_1y > 0 or ret_6m > 0) and ret_1m < 0 and ret_1w < 0)
    if strong:
        reversal_pen = -10
        reversal_flag = "🔴 Strong Reversal"
    elif mild:
        reversal_pen = -5
        reversal_flag = "⚠️ Mild Reversal"
    
    # Death Cross
    ma50 = calc_ma(df, base_idx, 50)
    ma200 = calc_ma(df, base_idx, 200)
    price = closes[base_idx]
    dc_pen = 0
    dc_flag = ""
    if ma50 is not None and ma200 is not None and ma50 < ma200:
        dc_pen = -5
        if price < ma50 and price < ma200:
            dc_flag = "💀💀 Death Cross + Below MAs"
        else:
            dc_flag = "💀 Death Cross"
    
    total = max(reversal_pen + dc_pen, -15)
    flags = " | ".join(f for f in [reversal_flag, dc_flag] if f)
    return {
        'reversal': reversal_pen, 'death_cross': dc_pen,
        'total': total, 'flags': flags,
        'ret_1y': ret_1y, 'ret_6m': ret_6m, 'ret_1m': ret_1m, 'ret_1w': ret_1w
    }

# ══════════════════════════════════════════════════════
# COMPUTE SCORES FOR BOTH BASE DATES
# ══════════════════════════════════════════════════════

def full_score(df, idx):
    ret_pctls = calc_return_percentiles(df, idx)
    vol_pctls = calc_volume_percentiles(df, idx)
    wp_ret = weighted_percentile(ret_pctls)
    wp_vol = weighted_percentile(vol_pctls)
    d1 = d1_score(wp_ret)
    d2 = d2_score(wp_vol)
    
    rsi = calc_rsi(df, idx)
    d3 = d3_score(rsi)
    
    price = df['Close'].values[idx]
    ma50 = calc_ma(df, idx, 50)
    ma200 = calc_ma(df, idx, 200)
    d4 = d4_score(price, ma50, ma200)
    
    vol = calc_volatility(df, idx)
    d5 = d5_score(vol)
    
    gross = d1 + d2 + d3 + d4 + d5
    penalties = calc_penalties(df, idx)
    net = gross + penalties['total']
    golden_cross = (ma50 is not None and ma200 is not None and ma50 > ma200)
    
    return {
        'date': df.iloc[idx]['Date'],
        'price': price,
        'ret_pctls': ret_pctls, 'vol_pctls': vol_pctls,
        'wp_ret': wp_ret, 'wp_vol': wp_vol,
        'd1': d1, 'd2': d2, 'd3': d3, 'd4': d4, 'd5': d5,
        'rsi': rsi, 'ma50': ma50, 'ma200': ma200,
        'golden_cross': golden_cross, 'volatility': vol,
        'gross': gross, 'penalties': penalties, 'net': net
    }

s1 = full_score(df, BD1_idx)
s2 = full_score(df, BD2_idx)

net_avg = (s1['net'] + s2['net']) / 2
gross_avg = (s1['gross'] + s2['gross']) / 2
delta = s2['net'] - s1['net']

def tier(score):
    if score >= 85: return "Very Strong ↑↑"
    if score >= 75: return "Strong ↑"
    if score >= 60: return "Moderate ↑"
    if score >= 45: return "Neutral →"
    if score >= 30: return "Weak ↓"
    return "Very Weak ↓↓"

momentum_tier = tier(net_avg)

print(f"\n{'='*50}")
print(f"Gold Momentum Score v2.2")
print(f"{'='*50}")
print(f"Net Score Avg: {net_avg:.2f}")
print(f"Tier: {momentum_tier}")
print(f"BD1 ({s1['date'].strftime('%Y-%m-%d')}): Net={s1['net']:.2f} Gross={s1['gross']:.2f}")
print(f"BD2 ({s2['date'].strftime('%Y-%m-%d')}): Net={s2['net']:.2f} Gross={s2['gross']:.2f}")
print(f"Delta: {delta:+.2f}")
print(f"Price: ${s2['price']:.1f}")
print(f"RSI: {s2['rsi']:.1f}")
print(f"Volatility: {s2['volatility']:.1f}%")
print(f"Penalties BD2: {s2['penalties']['total']} ({s2['penalties']['flags'] or 'None'})")

# ══════════════════════════════════════════════════════
# CSV OUTPUT
# ══════════════════════════════════════════════════════

csv_row = {
    'Rank': 1,
    'Ticker': 'GOLD',
    'Net_Score_Avg': round(net_avg, 2),
    'Gross_Score_Avg': round(gross_avg, 2),
    'Net_Score_BD1': round(s1['net'], 2),
    'Net_Score_BD2': round(s2['net'], 2),
    'Score_Delta': round(delta, 2),
    'Tier': momentum_tier,
    'D1_ReturnRank': round(s2['d1'], 2),
    'D2_VolumeRank': round(s2['d2'], 2),
    'D3_RSI': round(s2['d3'], 2),
    'D4_MA': round(s2['d4'], 2),
    'D5_Volatility': round(s2['d5'], 2),
    'WP_Return_Pct': round(s2['wp_ret'], 2),
    'WP_Volume_Pct': round(s2['wp_vol'], 2),
    'Ret_1Y_Pct': round(s2['penalties']['ret_1y'], 2),
    'Ret_6M_Pct': round(s2['penalties']['ret_6m'], 2),
    'Ret_3M_Pct': round(s2['ret_pctls']['3M']['return'], 2),
    'Ret_1M_Pct': round(s2['ret_pctls']['1M']['return'], 2),
    'Ret_1W_Pct': round(s2['ret_pctls']['1W']['return'], 2),
    'RSI_Value': round(s2['rsi'], 2),
    'MA50': round(s2['ma50'], 2) if s2['ma50'] else '',
    'MA200': round(s2['ma200'], 2) if s2['ma200'] else '',
    'Price': round(s2['price'], 2),
    'Golden_Cross': str(s2['golden_cross']),
    'Volatility_Pct': round(s2['volatility'], 2),
    'Penalty_Total': s2['penalties']['total'],
    'Penalty_Reversal': s2['penalties']['reversal'],
    'Penalty_DeathCross': s2['penalties']['death_cross'],
    'Warning_Flags': s2['penalties']['flags'] if s2['penalties']['flags'] else 'None',
    'News_Top20': 'FALSE',
    'Base_Date_1': s1['date'].strftime('%Y-%m-%d'),
    'Base_Date_2': s2['date'].strftime('%Y-%m-%d'),
    'As_Of_Running': AS_OF
}

csv_df = pd.DataFrame([csv_row])
base_dir = os.path.dirname(__file__)
csv_fixed = os.path.join(base_dir, 'output_momentum_gold.csv')
csv_ts = os.path.join(base_dir, f'output_momentum_gold_{TS_FILE}.csv')
csv_df.to_csv(csv_fixed, index=False, encoding='utf-8')
csv_df.to_csv(csv_ts, index=False, encoding='utf-8')
print(f"\nCSV saved: {csv_fixed}")
print(f"CSV saved: {csv_ts}")

# ══════════════════════════════════════════════════════
# EXCEL OUTPUT
# ══════════════════════════════════════════════════════

wb = Workbook()

# ── STYLES ──
gold_fill = PatternFill('solid', fgColor='FFD700')
green_fill = PatternFill('solid', fgColor='C6EFCE')
red_fill = PatternFill('solid', fgColor='FFC7CE')
blue_fill = PatternFill('solid', fgColor='BDD7EE')
gray_fill = PatternFill('solid', fgColor='D9D9D9')
dark_fill = PatternFill('solid', fgColor='333333')
header_font = Font(bold=True, size=11, color='FFFFFF')
title_font = Font(bold=True, size=14, color='333333')
big_font = Font(bold=True, size=18, color='1F4E79')
val_font = Font(size=11)
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header_row(ws, row, cols, fill=None):
    f = fill or PatternFill('solid', fgColor='1F4E79')
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = f
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

def style_cell(ws, row, col, fmt=None):
    cell = ws.cell(row=row, column=col)
    cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center')
    if fmt: cell.number_format = fmt
    return cell

# ════════════ SHEET 1: SUMMARY ════════════
ws1 = wb.active
ws1.title = "Gold Momentum v2.2"
ws1.sheet_properties.tabColor = "FFD700"

# Title
ws1.merge_cells('A1:H1')
ws1['A1'] = "🥇 Gold Momentum Score v2.2"
ws1['A1'].font = big_font
ws1['A1'].alignment = Alignment(horizontal='center')

ws1.merge_cells('A2:H2')
ws1['A2'] = f"Run: {AS_OF}  |  BD1: {s1['date'].strftime('%Y-%m-%d')}  |  BD2: {s2['date'].strftime('%Y-%m-%d')}"
ws1['A2'].font = Font(size=10, color='666666')
ws1['A2'].alignment = Alignment(horizontal='center')

# Score hero
ws1.merge_cells('A4:B4')
ws1['A4'] = "Net Score Avg"
ws1['A4'].font = Font(bold=True, size=12)
ws1['C4'] = round(net_avg, 2)
ws1['C4'].font = Font(bold=True, size=16, color='1F4E79')
ws1['C4'].number_format = '0.00'

ws1['D4'] = momentum_tier
ws1['D4'].font = Font(bold=True, size=14)
ws1['D4'].fill = gold_fill

ws1['F4'] = f"Price: ${s2['price']:,.1f}"
ws1['F4'].font = Font(bold=True, size=12)

# Summary table
row = 6
headers = ['Metric', 'BD1', 'BD2', 'Average', 'Delta']
for i, h in enumerate(headers):
    ws1.cell(row=row, column=i+1, value=h)
style_header_row(ws1, row, len(headers))

data_rows = [
    ['Net Score', s1['net'], s2['net'], net_avg, delta],
    ['Gross Score', s1['gross'], s2['gross'], gross_avg, s2['gross']-s1['gross']],
    ['D1 Return Rank', s1['d1'], s2['d1'], (s1['d1']+s2['d1'])/2, s2['d1']-s1['d1']],
    ['D2 Volume Rank', s1['d2'], s2['d2'], (s1['d2']+s2['d2'])/2, s2['d2']-s1['d2']],
    ['D3 RSI', s1['d3'], s2['d3'], (s1['d3']+s2['d3'])/2, s2['d3']-s1['d3']],
    ['D4 MA Trend', s1['d4'], s2['d4'], (s1['d4']+s2['d4'])/2, s2['d4']-s1['d4']],
    ['D5 Volatility', s1['d5'], s2['d5'], (s1['d5']+s2['d5'])/2, s2['d5']-s1['d5']],
    ['Penalty', s1['penalties']['total'], s2['penalties']['total'], (s1['penalties']['total']+s2['penalties']['total'])/2, s2['penalties']['total']-s1['penalties']['total']],
]
for r, row_data in enumerate(data_rows):
    rn = row + 1 + r
    for c, val in enumerate(row_data):
        cell = style_cell(ws1, rn, c+1, '0.00' if isinstance(val, float) else None)
        cell.value = val
        if c == 4 and isinstance(val, (int, float)):
            cell.font = Font(color='006100' if val >= 0 else '9C0006')

# Return performance
ret_row = row + 1 + len(data_rows) + 2
ws1.cell(row=ret_row, column=1, value="Return Performance (BD2)").font = Font(bold=True, size=12)
ret_row += 1
ret_headers = ['Period', 'Days', 'Return %', 'Rolling Percentile', 'Weight']
for i, h in enumerate(ret_headers):
    ws1.cell(row=ret_row, column=i+1, value=h)
style_header_row(ws1, ret_row, len(ret_headers))

for p in WEIGHT_ORDER:
    ret_row += 1
    vals = [p, LOOKBACK[p], s2['ret_pctls'][p]['return'], s2['ret_pctls'][p]['percentile'], WEIGHTS[p]*100]
    for c, v in enumerate(vals):
        cell = style_cell(ws1, ret_row, c+1, '0.00')
        cell.value = round(v, 2) if isinstance(v, float) else v

# Column widths
for c in range(1, 9):
    ws1.column_dimensions[get_column_letter(c)].width = 18

# ════════════ SHEET 2: RETURN PERCENTILE DETAIL ════════════
ws2 = wb.create_sheet("Return Percentile Detail")
ws2.sheet_properties.tabColor = "4472C4"

ws2['A1'] = "📊 Return Percentile Detail (Rolling 252-day)"
ws2['A1'].font = title_font
ws2.merge_cells('A1:F1')

for bd_label, sc, start_row in [("Base Date 1", s1, 3), ("Base Date 2", s2, 10)]:
    ws2.cell(row=start_row, column=1, value=f"{bd_label}: {sc['date'].strftime('%Y-%m-%d')}").font = Font(bold=True, size=11)
    hr = start_row + 1
    cols = ['Period', 'Days', 'Weight', 'Raw Return %', 'Percentile Rank', 'Weighted Contribution']
    for i, h in enumerate(cols):
        ws2.cell(row=hr, column=i+1, value=h)
    style_header_row(ws2, hr, len(cols))
    
    total_wp = 0
    for p in WEIGHT_ORDER:
        hr += 1
        pdata = sc['ret_pctls'][p]
        contrib = pdata['percentile'] * WEIGHTS[p]
        total_wp += contrib
        vals = [p, LOOKBACK[p], f"{WEIGHTS[p]*100:.0f}%", round(pdata['return'], 2), round(pdata['percentile'], 2), round(contrib, 2)]
        for c, v in enumerate(vals):
            style_cell(ws2, hr, c+1).value = v
    
    hr += 1
    ws2.cell(row=hr, column=1, value="Total WP_Return").font = Font(bold=True)
    ws2.cell(row=hr, column=5, value=round(sc['wp_ret'], 2)).font = Font(bold=True)
    ws2.cell(row=hr, column=6, value=round(sc['wp_ret'], 2)).font = Font(bold=True)
    hr += 1
    ws2.cell(row=hr, column=1, value="D1 Score (WP/100 × 20)").font = Font(bold=True)
    ws2.cell(row=hr, column=5, value=round(sc['d1'], 2)).font = Font(bold=True, color='1F4E79')

for c in range(1, 7):
    ws2.column_dimensions[get_column_letter(c)].width = 22

# ════════════ SHEET 3: VOLUME PERCENTILE DETAIL ════════════
ws3 = wb.create_sheet("Volume Percentile Detail")
ws3.sheet_properties.tabColor = "70AD47"

ws3['A1'] = "📊 Volume Percentile Detail (Rolling 252-day)"
ws3['A1'].font = title_font
ws3.merge_cells('A1:F1')

for bd_label, sc, start_row in [("Base Date 1", s1, 3), ("Base Date 2", s2, 10)]:
    ws3.cell(row=start_row, column=1, value=f"{bd_label}: {sc['date'].strftime('%Y-%m-%d')}").font = Font(bold=True)
    hr = start_row + 1
    cols = ['Period', 'Days', 'Weight', 'Cum Volume', 'Percentile Rank', 'Weighted Contribution']
    for i, h in enumerate(cols):
        ws3.cell(row=hr, column=i+1, value=h)
    style_header_row(ws3, hr, len(cols))
    
    for p in WEIGHT_ORDER:
        hr += 1
        pdata = sc['vol_pctls'][p]
        contrib = pdata['percentile'] * WEIGHTS[p]
        vals = [p, LOOKBACK[p], f"{WEIGHTS[p]*100:.0f}%", round(pdata['volume'], 0), round(pdata['percentile'], 2), round(contrib, 2)]
        for c, v in enumerate(vals):
            style_cell(ws3, hr, c+1).value = v
    
    hr += 1
    ws3.cell(row=hr, column=1, value="Total WP_Volume").font = Font(bold=True)
    ws3.cell(row=hr, column=5, value=round(sc['wp_vol'], 2)).font = Font(bold=True)
    hr += 1
    ws3.cell(row=hr, column=1, value="D2 Score (WP/100 × 20)").font = Font(bold=True)
    ws3.cell(row=hr, column=5, value=round(sc['d2'], 2)).font = Font(bold=True, color='1F4E79')

for c in range(1, 7):
    ws3.column_dimensions[get_column_letter(c)].width = 22

# ════════════ SHEET 4: WARNING FLAGS ════════════
ws4 = wb.create_sheet("🚨 Warning Flags")
ws4.sheet_properties.tabColor = "FF0000"

ws4['A1'] = "🚨 Warning Flags & Penalty System"
ws4['A1'].font = title_font
ws4.merge_cells('A1:F1')

for bd_label, sc, start_row in [("Base Date 2 (Latest)", s2, 3), ("Base Date 1", s1, 12)]:
    ws4.cell(row=start_row, column=1, value=f"{bd_label}: {sc['date'].strftime('%Y-%m-%d')}").font = Font(bold=True, size=11)
    hr = start_row + 1
    cols = ['Penalty Type', 'Pts', 'Condition', 'Current Values', 'Status']
    for i, h in enumerate(cols):
        ws4.cell(row=hr, column=i+1, value=h)
    style_header_row(ws4, hr, len(cols), PatternFill('solid', fgColor='C00000'))
    
    pen = sc['penalties']
    # Reversal
    hr += 1
    rev_status = "TRIGGERED" if pen['reversal'] != 0 else "CLEAN"
    vals = ['Momentum Reversal', pen['reversal'],
            '1Y>0 & 1M<0 & 1W<0 (mild) or 1Y>20 & 1M<-5 & 1W<-3 (strong)',
            f"1Y={pen['ret_1y']:.1f}% 6M={pen['ret_6m']:.1f}% 1M={pen['ret_1m']:.1f}% 1W={pen['ret_1w']:.1f}%",
            rev_status]
    for c, v in enumerate(vals):
        cell = style_cell(ws4, hr, c+1)
        cell.value = v
        if c == 4:
            cell.fill = red_fill if rev_status == "TRIGGERED" else green_fill
    
    # Death Cross
    hr += 1
    dc_status = "TRIGGERED" if pen['death_cross'] != 0 else "CLEAN"
    ma50_v = sc['ma50'] if sc['ma50'] else 0
    ma200_v = sc['ma200'] if sc['ma200'] else 0
    vals = ['Death Cross / MA Breakdown', pen['death_cross'],
            'MA50 < MA200',
            f"MA50={ma50_v:.1f} MA200={ma200_v:.1f} Price={sc['price']:.1f}",
            dc_status]
    for c, v in enumerate(vals):
        cell = style_cell(ws4, hr, c+1)
        cell.value = v
        if c == 4:
            cell.fill = red_fill if dc_status == "TRIGGERED" else green_fill
    
    # Total
    hr += 2
    ws4.cell(row=hr, column=1, value="Total Penalty").font = Font(bold=True, size=12)
    ws4.cell(row=hr, column=2, value=pen['total']).font = Font(bold=True, size=12, color='9C0006' if pen['total'] < 0 else '006100')
    ws4.cell(row=hr, column=3, value=f"Flags: {pen['flags'] or 'None'}").font = Font(italic=True)
    hr += 1
    ws4.cell(row=hr, column=1, value=f"Gross Score: {sc['gross']:.2f}  →  Net Score: {sc['net']:.2f}").font = Font(bold=True)

for c in range(1, 6):
    ws4.column_dimensions[get_column_letter(c)].width = 35

# ════════════ SHEET 5: METHODOLOGY ════════════
ws5 = wb.create_sheet("Methodology v2.2")
ws5.sheet_properties.tabColor = "7030A0"

ws5['A1'] = "📘 Methodology — Gold Momentum Scoring v2.2"
ws5['A1'].font = title_font
ws5.merge_cells('A1:C1')

methods = [
    ("D1: Return Rank (0-20)", "Rolling Percentile of returns vs self over 252 days\nWeights: 1Y=30%, 6M=25%, 3M=20%, 1M=15%, 1W=10%\nScore = Weighted Percentile / 100 × 20"),
    ("D2: Volume Rank (0-20)", "Rolling Percentile of cumulative volume vs self over 252 days\nSame weights as D1\nScore = Weighted Percentile / 100 × 20"),
    ("D3: RSI (0-20)", "14-day RSI\n50-70→20pts, 40-49→16pts, 71-80→14pts\n30-39→12pts, >80→10pts, <30→6pts"),
    ("D4: MA Trend (0-20)", "Price>MA50→+7, Price>MA200→+7, Golden Cross(MA50>MA200)→+6\nMax: 20 pts"),
    ("D5: Volatility (0-20)", "21-day annualized volatility\n≤20%→20, 21-30%→18, 31-40%→14, 41-50%→11\n51-60%→8, 61-80%→5, >80%→2"),
    ("Penalty System", "Mild Reversal: -5 (1Y>0 & 1M<0 & 1W<0)\nStrong Reversal: -10 (1Y>20% & 1M<-5% & 1W<-3%)\nDeath Cross: -5 (MA50<MA200)\nMax penalty: -15 (capped)"),
    ("Rolling Percentile (vs SP500)", "SP500 uses cross-sectional percentile (rank among 500 stocks)\nGold uses Rolling Percentile (rank vs own 252-day history)\nThis adapts the same framework for a single asset"),
    ("Data Source", f"Yahoo Finance GC=F (Gold Futures COMEX)\nRepo: github.com/jptrustlearning/gold\nFile: gold_prices.csv\nRange: {df.iloc[0]['Date'].strftime('%Y-%m-%d')} to {df.iloc[-1]['Date'].strftime('%Y-%m-%d')}")
]

row = 3
for title, desc in methods:
    ws5.cell(row=row, column=1, value=title).font = Font(bold=True, size=11, color='7030A0')
    ws5.cell(row=row, column=2, value=desc).alignment = Alignment(wrap_text=True)
    row += 3

ws5.column_dimensions['A'].width = 30
ws5.column_dimensions['B'].width = 70

# ── SAVE EXCEL ──
excel_path = os.path.join(base_dir, 'Gold_Momentum_v2.2.xlsx')
wb.save(excel_path)
print(f"\nExcel saved: {excel_path}")

print(f"\n✅ All outputs generated successfully!")
print(f"   CSV files: output_momentum_gold.csv + output_momentum_gold_{TS_FILE}.csv")
print(f"   Excel: Gold_Momentum_v2.2.xlsx")
