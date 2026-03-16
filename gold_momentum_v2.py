#!/usr/bin/env python3
"""
Gold Momentum Scoring System v3.0 (100-Scale Edition)
JP Trust Learning

V3 changes scoring scale:
- All dimensions D1-D6 now scored 0-100 each
- Total score = average of all 6 dimensions (0-100)
- Penalty scaled proportionally from old system
- Net Score = Gross (avg D1-D6) + Penalty_Scaled
- Net Score range: ~-14 to 100
"""

import pandas as pd
import numpy as np
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, sys

# ── CONFIG ──
ROLLING_WINDOW = 252
LOOKBACK = {'1W': 5, '1M': 21, '3M': 63, '6M': 126, '1Y': 252}
WEIGHTS = {'1Y': 0.30, '6M': 0.25, '3M': 0.20, '1M': 0.15, '1W': 0.10}
WEIGHT_ORDER = ['1Y', '6M', '3M', '1M', '1W']

RUN_TS = datetime.now(timezone.utc)
AS_OF = RUN_TS.strftime("%d/%m/%Y %H:%M UTC")
TS_FILE = RUN_TS.strftime("%d%m%Y_%H%M")

# ── LOAD DATA ──
base_dir = os.path.dirname(os.path.abspath(__file__))

def load_price_csv(filename):
    path = os.path.join(base_dir, filename)
    if not os.path.exists(path):
        print(f"⚠️ {filename} not found — skipping")
        return None
    df = pd.read_csv(path, encoding='utf-8-sig')
    df.columns = ['Date', 'Open', 'High', 'Low', 'Close', 'Volume']
    df['Date'] = pd.to_datetime(df['Date'])
    df = df.sort_values('Date').reset_index(drop=True)
    return df

df = load_price_csv('gold_prices.csv')
df_dxy = load_price_csv('dxy_prices.csv')
df_vix = load_price_csv('vix_prices.csv')

if df is None:
    print("❌ gold_prices.csv not found — cannot continue")
    sys.exit(1)

# ── BASE DATES ──
BD2_idx = len(df) - 1
BD1_idx = len(df) - 6
BD1_date = df.iloc[BD1_idx]['Date']
BD2_date = df.iloc[BD2_idx]['Date']

print(f"Gold Momentum Scoring v3.0 (100-Scale Edition)")
print(f"{'='*55}")
print(f"Base Date 1: {BD1_date.strftime('%Y-%m-%d')} (idx={BD1_idx})")
print(f"Base Date 2: {BD2_date.strftime('%Y-%m-%d')} (idx={BD2_idx})")
print(f"Total gold rows: {len(df)}")
if df_dxy is not None:
    print(f"DXY rows: {len(df_dxy)} (latest: {df_dxy['Date'].max().strftime('%Y-%m-%d')})")
if df_vix is not None:
    print(f"VIX rows: {len(df_vix)} (latest: {df_vix['Date'].max().strftime('%Y-%m-%d')})")

# ══════════════════════════════════════════════════════
# ORIGINAL 5 DIMENSIONS (unchanged from v2.2)
# ══════════════════════════════════════════════════════

def compute_return(closes, end_idx, period_days):
    start_idx = end_idx - period_days
    if start_idx < 0:
        return None
    return (closes[end_idx] - closes[start_idx]) / closes[start_idx] * 100

def rolling_percentile(series_values, current_val, window=ROLLING_WINDOW):
    valid = series_values[~np.isnan(series_values)]
    if len(valid) < 10:
        return 50.0
    count_below = np.sum(valid < current_val)
    return count_below / (len(valid) - 1) * 100 if len(valid) > 1 else 50.0

def calc_return_percentiles(df, base_idx):
    closes = df['Close'].values
    results = {}
    for period, days in LOOKBACK.items():
        current_ret = compute_return(closes, base_idx, days)
        if current_ret is None:
            results[period] = {'return': 0, 'percentile': 50}
            continue
        rolling_rets = []
        start = max(0, base_idx - ROLLING_WINDOW)
        for i in range(start, base_idx):
            r = compute_return(closes, i, days)
            if r is not None:
                rolling_rets.append(r)
        pctl = rolling_percentile(np.array(rolling_rets), current_ret) if rolling_rets else 50
        results[period] = {'return': current_ret, 'percentile': pctl}
    return results

def calc_volume_percentiles(df, base_idx):
    volumes = df['Volume'].values
    results = {}
    for period, days in LOOKBACK.items():
        end = base_idx + 1
        start = end - days
        if start < 0:
            results[period] = {'volume': 0, 'percentile': 50}
            continue
        current_vol = np.sum(volumes[start:end])
        rolling_vols = []
        roll_start = max(0, base_idx - ROLLING_WINDOW)
        for i in range(roll_start, base_idx):
            s = i + 1 - days
            if s < 0:
                continue
            rolling_vols.append(np.sum(volumes[s:i+1]))
        pctl = rolling_percentile(np.array(rolling_vols), current_vol) if rolling_vols else 50
        results[period] = {'volume': current_vol, 'percentile': pctl}
    return results

def weighted_percentile(pctl_dict):
    return sum(pctl_dict[p]['percentile'] * WEIGHTS[p] for p in WEIGHT_ORDER)

def d1_score(wp): return wp  # WP is already 0-100
def d2_score(wp): return wp  # WP is already 0-100

def calc_rsi(df, base_idx, period=14):
    start = base_idx - 29
    if start < 0: start = 0
    closes = df['Close'].values[start:base_idx+1]
    if len(closes) < 2:
        return 50
    deltas = np.diff(closes)
    gains = np.where(deltas > 0, deltas, 0)
    losses = np.where(deltas < 0, -deltas, 0)
    last_n = min(period, len(gains))
    avg_gain = np.mean(gains[-last_n:])
    avg_loss = np.mean(losses[-last_n:])
    if avg_loss == 0:
        return 100.0
    rs = avg_gain / avg_loss
    return 100 - (100 / (1 + rs))

def d3_score(rsi):
    if 50 <= rsi <= 70: return 100
    if 40 <= rsi < 50: return 80
    if 70 < rsi <= 80: return 70
    if 30 <= rsi < 40: return 60
    if rsi > 80: return 50
    return 30

def calc_ma(df, base_idx, window):
    start = base_idx + 1 - window
    if start < 0: return None
    return np.mean(df['Close'].values[start:base_idx+1])

def d4_score(price, ma50, ma200):
    pts = 0
    if ma50 is not None and price > ma50: pts += 35
    if ma200 is not None and price > ma200: pts += 35
    if ma50 is not None and ma200 is not None and ma50 > ma200: pts += 30
    return min(pts, 100)

def calc_volatility(df, base_idx):
    """Directional Volatility — แยก upside/downside vol แล้วคำนวณ ratio"""
    start = base_idx - 20
    if start < 0: start = 0
    closes = df['Close'].values[start:base_idx+1]
    if len(closes) < 2:
        return {'abs_vol': 0, 'up_vol': 0, 'down_vol': 0, 'vol_ratio': 1.0}
    rets = np.diff(closes) / closes[:-1]
    abs_vol = np.std(rets) * np.sqrt(252) * 100

    up_rets = rets[rets > 0]
    down_rets = rets[rets < 0]

    up_vol = np.std(up_rets) * np.sqrt(252) * 100 if len(up_rets) >= 2 else 0
    down_vol = np.std(down_rets) * np.sqrt(252) * 100 if len(down_rets) >= 2 else 0

    # vol_ratio = down_vol / up_vol — สูง = downside dominant
    if up_vol > 0:
        vol_ratio = down_vol / up_vol
    elif down_vol > 0:
        vol_ratio = 999  # ไม่มี up days เลย → extreme downside
    else:
        vol_ratio = 1.0  # ไม่มีทั้งสองฝั่ง → neutral

    return {'abs_vol': abs_vol, 'up_vol': up_vol, 'down_vol': down_vol, 'vol_ratio': vol_ratio}

def d5_score(vol_data):
    """D5 Directional Volatility Score (0-100) — ใช้ vol_ratio"""
    ratio = vol_data['vol_ratio'] if isinstance(vol_data, dict) else 1.0
    if ratio <= 0.6:  return 100   # upside dominant มาก
    if ratio <= 0.8:  return 85
    if ratio <= 1.0:  return 70    # balanced
    if ratio <= 1.2:  return 55
    if ratio <= 1.5:  return 40    # downside เริ่มครอบงำ
    if ratio <= 2.0:  return 20
    return 10                       # downside dominant

def calc_penalties(df, base_idx):
    closes = df['Close'].values
    ret_1y = compute_return(closes, base_idx, 252) or 0
    ret_6m = compute_return(closes, base_idx, 126) or 0
    ret_1m = compute_return(closes, base_idx, 21) or 0
    ret_1w = compute_return(closes, base_idx, 5) or 0

    reversal_pen = 0
    reversal_flag = ""
    strong = (ret_1y > 20 and ret_1m < -5 and ret_1w < -3)
    mild = ((ret_1y > 0 or ret_6m > 0) and ret_1m < 0 and ret_1w < 0)
    if strong:
        reversal_pen = -10
        reversal_flag = "🔴 Strong Reversal"
    elif mild:
        reversal_pen = -5
        reversal_flag = "⚠️ Mild Reversal"

    ma50 = calc_ma(df, base_idx, 50)
    ma200 = calc_ma(df, base_idx, 200)
    price = closes[base_idx]
    dc_pen = 0
    dc_flag = ""
    if ma50 is not None and ma200 is not None and ma50 < ma200:
        dc_pen = -5
        if price < ma50 and price < ma200:
            dc_flag = "💀💀 Death Cross + Below MAs"
        else:
            dc_flag = "💀 Death Cross"

    total = max(reversal_pen + dc_pen, -15)
    flags = " | ".join(f for f in [reversal_flag, dc_flag] if f)
    return {
        'reversal': reversal_pen, 'death_cross': dc_pen,
        'total': total, 'flags': flags,
        'ret_1y': ret_1y, 'ret_6m': ret_6m, 'ret_1m': ret_1m, 'ret_1w': ret_1w
    }

# ══════════════════════════════════════════════════════
# NEW: DIMENSION 6 — EXTERNAL CONTEXT (DXY + VIX)
# ══════════════════════════════════════════════════════

def find_closest_idx(ext_df, target_date, max_gap_days=5):
    """Find the closest date index in external data to the target date."""
    if ext_df is None:
        return None
    diffs = (ext_df['Date'] - target_date).abs()
    min_diff = diffs.min()
    if min_diff.days > max_gap_days:
        return None
    return diffs.idxmin()

def calc_external_return(ext_df, end_idx, period_days):
    """Calculate return for external asset."""
    if ext_df is None or end_idx is None:
        return None
    start_idx = end_idx - period_days
    if start_idx < 0:
        return None
    return (ext_df['Close'].values[end_idx] - ext_df['Close'].values[start_idx]) / ext_df['Close'].values[start_idx] * 100

def calc_d6_external(df_gold, gold_idx, df_dxy, df_vix):
    """
    Dimension 6: External Context Score (±10 pts total)
    
    Part A — DXY Divergence (±5 pts):
      Gold up + DXY up (divergence)   → +5 (gold rising DESPITE strong dollar = very bullish)
      Gold up + DXY down (normal)     → +2 (gold rising with weak dollar = expected)
      Gold down + DXY down            →  0 (neutral)
      Gold down + DXY up              → -5 (headwind — dollar strength dragging gold)
      DXY data unavailable            →  0
    
    Part B — VIX Regime (±5 pts):
      VIX > 30 + Gold up              → +5 (safe-haven demand confirmed)
      VIX 20-30 + Gold up             → +3 (elevated fear, gold benefiting)
      VIX < 20 + Gold up              → +1 (calm market, gold rising on own merit)
      VIX > 30 + Gold down            → -3 (panic selling even gold)
      VIX 20-30 + Gold down           → -2 (moderate fear, gold not benefiting)
      VIX < 20 + Gold down            →  0 (calm market, gold drifting — neutral)
      VIX data unavailable            →  0
    
    Total D6 range: -10 to +10
    """
    gold_date = df_gold.iloc[gold_idx]['Date']
    gold_closes = df_gold['Close'].values
    gold_1m = compute_return(gold_closes, gold_idx, 21)
    if gold_1m is None:
        gold_1m = 0
    gold_up = gold_1m >= 0

    # ── Part A: DXY Divergence ──
    dxy_score = 0
    dxy_1m = None
    dxy_signal = "N/A"
    
    if df_dxy is not None:
        dxy_idx = find_closest_idx(df_dxy, gold_date)
        if dxy_idx is not None:
            dxy_1m = calc_external_return(df_dxy, dxy_idx, 21)
            if dxy_1m is not None:
                dxy_up = dxy_1m > 0
                if gold_up and dxy_up:
                    dxy_score = +5
                    dxy_signal = "🟢 Bullish Divergence (gold up despite strong $)"
                elif gold_up and not dxy_up:
                    dxy_score = +2
                    dxy_signal = "🔵 Normal (gold up + weak $)"
                elif not gold_up and not dxy_up:
                    dxy_score = 0
                    dxy_signal = "⚪ Neutral (both down)"
                else:  # gold down, dxy up
                    dxy_score = -5
                    dxy_signal = "🔴 Headwind (gold down + strong $)"

    # ── Part B: VIX Regime ──
    vix_score = 0
    vix_level = None
    vix_signal = "N/A"
    
    if df_vix is not None:
        vix_idx = find_closest_idx(df_vix, gold_date)
        if vix_idx is not None:
            vix_level = df_vix['Close'].values[vix_idx]
            if gold_up:
                if vix_level > 30:
                    vix_score = +5
                    vix_signal = "🟢 Safe-Haven Confirmed (VIX>30 + gold up)"
                elif vix_level >= 20:
                    vix_score = +3
                    vix_signal = "🔵 Elevated Fear (VIX 20-30 + gold up)"
                else:
                    vix_score = +1
                    vix_signal = "⚪ Calm Rally (VIX<20 + gold up)"
            else:
                if vix_level > 30:
                    vix_score = -3
                    vix_signal = "🔴 Panic Selling (VIX>30 + gold down)"
                elif vix_level >= 20:
                    vix_score = -2
                    vix_signal = "🟠 Fear Not Helping (VIX 20-30 + gold down)"
                else:
                    vix_score = 0
                    vix_signal = "⚪ Calm Drift (VIX<20 + gold down)"

    total_d6 = max(min(dxy_score + vix_score, 10), -10)
    # Scale from ±10 to 0-100: -10→0, 0→50, +10→100
    d6_scaled = (total_d6 + 10) / 20 * 100
    
    return {
        'd6_total': total_d6,       # raw ±10 (for signals/display)
        'd6_scaled': d6_scaled,      # 0-100 (for scoring)
        'dxy_score': dxy_score,
        'vix_score': vix_score,
        'dxy_1m': dxy_1m,
        'vix_level': vix_level,
        'dxy_signal': dxy_signal,
        'vix_signal': vix_signal,
        'gold_1m': gold_1m
    }


# ══════════════════════════════════════════════════════
# COMPUTE FULL SCORES
# ══════════════════════════════════════════════════════

def full_score(df, idx, df_dxy, df_vix):
    ret_pctls = calc_return_percentiles(df, idx)
    vol_pctls = calc_volume_percentiles(df, idx)
    wp_ret = weighted_percentile(ret_pctls)
    wp_vol = weighted_percentile(vol_pctls)
    d1 = d1_score(wp_ret)
    d2 = d2_score(wp_vol)
    
    rsi = calc_rsi(df, idx)
    d3 = d3_score(rsi)
    
    price = df['Close'].values[idx]
    ma50 = calc_ma(df, idx, 50)
    ma200 = calc_ma(df, idx, 200)
    d4 = d4_score(price, ma50, ma200)
    
    vol_data = calc_volatility(df, idx)
    d5 = d5_score(vol_data)
    vol = vol_data['abs_vol']  # backward compat — absolute vol for display
    
    gross = d1 + d2 + d3 + d4 + d5
    penalties = calc_penalties(df, idx)
    
    # NEW: External context
    ext = calc_d6_external(df, idx, df_dxy, df_vix)
    d6 = ext['d6_scaled']  # 0-100 scale
    
    # V3 Scoring: each dimension is 0-100, gross = average of all 6
    gross_avg_dims = (d1 + d2 + d3 + d4 + d5 + d6) / 6
    
    # Penalty: scale from old system (-15 on ~110) to new system (/100)
    # Keep same proportional impact: -15/110 ≈ -13.6 on /100 scale
    penalty_scaled = penalties['total'] * (100 / 110)
    
    # Net Score = Gross Average - Penalty (on 0-100 scale)
    net = gross_avg_dims + penalty_scaled
    golden_cross = (ma50 is not None and ma200 is not None and ma50 > ma200)
    
    return {
        'date': df.iloc[idx]['Date'],
        'price': price,
        'ret_pctls': ret_pctls, 'vol_pctls': vol_pctls,
        'wp_ret': wp_ret, 'wp_vol': wp_vol,
        'd1': d1, 'd2': d2, 'd3': d3, 'd4': d4, 'd5': d5, 'd6': d6,
        'd6_raw': ext['d6_total'],  # raw ±10 for display
        'rsi': rsi, 'ma50': ma50, 'ma200': ma200,
        'golden_cross': golden_cross, 'volatility': vol,
        'vol_ratio': vol_data['vol_ratio'],
        'up_vol': vol_data['up_vol'], 'down_vol': vol_data['down_vol'],
        'gross': gross_avg_dims, 'penalties': penalties,
        'penalty_scaled': penalty_scaled,
        'net': net,
        'external': ext
    }

s1 = full_score(df, BD1_idx, df_dxy, df_vix)
s2 = full_score(df, BD2_idx, df_dxy, df_vix)

net_avg = (s1['net'] + s2['net']) / 2
gross_avg = (s1['gross'] + s2['gross']) / 2
delta = s2['net'] - s1['net']

def tier(score):
    # Tier uses clamped score (0-100 range) for consistency
    clamped = max(0, min(100, score))
    if clamped >= 85: return "Very Strong ↑↑"
    if clamped >= 75: return "Strong ↑"
    if clamped >= 60: return "Moderate ↑"
    if clamped >= 45: return "Neutral →"
    if clamped >= 30: return "Weak ↓"
    return "Very Weak ↓↓"

momentum_tier = tier(net_avg)

print(f"\n{'='*55}")
print(f"Gold Momentum Score v3.0 (100-Scale Edition)")
print(f"{'='*55}")
print(f"Net Score Avg:  {net_avg:.2f}  ({momentum_tier})")
print(f"Gross Score Avg: {gross_avg:.2f}")
print(f"BD1 ({s1['date'].strftime('%Y-%m-%d')}): Net={s1['net']:.2f}  D6={s1['d6']:.1f}/100 (raw {s1['d6_raw']:+d})")
print(f"BD2 ({s2['date'].strftime('%Y-%m-%d')}): Net={s2['net']:.2f}  D6={s2['d6']:.1f}/100 (raw {s2['d6_raw']:+d})")
print(f"Delta: {delta:+.2f}")
print(f"Price: ${s2['price']:.1f}")
print(f"RSI: {s2['rsi']:.1f} | Volatility: {s2['volatility']:.1f}% | Vol Ratio (D/U): {s2['vol_ratio']:.2f}")
print(f"Penalties: {s2['penalties']['total']} (scaled: {s2['penalty_scaled']:.1f}) ({s2['penalties']['flags'] or 'None'})")
print(f"\n── External Context (BD2) ──")
print(f"D6 Raw: {s2['d6_raw']:+d} → Scaled: {s2['d6']:.1f}/100")
print(f"  DXY: {s2['external']['dxy_score']:+d}  ({s2['external']['dxy_signal']})")
print(f"  VIX: {s2['external']['vix_score']:+d}  ({s2['external']['vix_signal']})")
if s2['external']['dxy_1m'] is not None:
    print(f"  DXY 1M Return: {s2['external']['dxy_1m']:.2f}%")
if s2['external']['vix_level'] is not None:
    print(f"  VIX Level: {s2['external']['vix_level']:.2f}")

# ══════════════════════════════════════════════════════
# MULTI-TF PIVOT POINTS + CONFLUENCE ZONES
# ══════════════════════════════════════════════════════

def calc_pivot_levels(high, low, close):
    """Classic Pivot Points: PP, R1-R3, S1-S3."""
    PP = (high + low + close) / 3
    R1 = 2 * PP - low
    S1 = 2 * PP - high
    R2 = PP + (high - low)
    S2 = PP - (high - low)
    R3 = high + 2 * (PP - low)
    S3 = low - 2 * (high - PP)
    return {'PP': round(PP, 2), 'R1': round(R1, 2), 'R2': round(R2, 2), 'R3': round(R3, 2),
            'S1': round(S1, 2), 'S2': round(S2, 2), 'S3': round(S3, 2)}

def calc_multi_tf_pivots(df):
    """
    Calculate pivot points for D1, W1, and MN timeframes from daily OHLCV data.
    
    - D1: uses previous completed trading day's H/L/C
    - W1: uses previous completed week's aggregated H/L/C
    - MN: uses previous completed month's aggregated H/L/C
    
    Returns dict with 'D1', 'W1', 'MN' keys (each containing pivot levels + source H/L/C)
    """
    df = df.copy()
    df['Date'] = pd.to_datetime(df['Date'])
    df = df.sort_values('Date').reset_index(drop=True)
    
    result = {}
    current_price = df['Close'].values[-1]
    latest_date = df['Date'].max()
    
    # ── D1: Last completed day H/L/C (วันล่าสุดที่มีราคาปิดแล้ว) ──
    if len(df) >= 1:
        prev = df.iloc[-1]
        levels = calc_pivot_levels(prev['High'], prev['Low'], prev['Close'])
        result['D1'] = {**levels, 'H': round(prev['High'], 2), 'L': round(prev['Low'], 2),
                        'C': round(prev['Close'], 2), 'date': prev['Date'].strftime('%Y-%m-%d')}
    
    # ── W1: Previous completed week H/L/C ──
    df['iso_year'] = df['Date'].dt.isocalendar().year.astype(int)
    df['iso_week'] = df['Date'].dt.isocalendar().week.astype(int)
    df['yw_key'] = df['iso_year'] * 100 + df['iso_week']
    
    current_yw = latest_date.isocalendar()
    current_yw_key = current_yw.year * 100 + current_yw.week
    
    # Filter out current (incomplete) week
    weekly = df[df['yw_key'] < current_yw_key].groupby('yw_key').agg(
        High=('High', 'max'), Low=('Low', 'min'), Close=('Close', 'last'),
        Date_last=('Date', 'max')
    ).reset_index().sort_values('yw_key')
    
    if len(weekly) >= 1:
        prev_w = weekly.iloc[-1]
        levels = calc_pivot_levels(prev_w['High'], prev_w['Low'], prev_w['Close'])
        result['W1'] = {**levels, 'H': round(prev_w['High'], 2), 'L': round(prev_w['Low'], 2),
                        'C': round(prev_w['Close'], 2), 'date': prev_w['Date_last'].strftime('%Y-%m-%d')}
    
    # ── MN: Previous completed month H/L/C ──
    current_ym = latest_date.strftime('%Y-%m')
    df['ym'] = df['Date'].dt.strftime('%Y-%m')
    
    monthly = df[df['ym'] < current_ym].groupby('ym').agg(
        High=('High', 'max'), Low=('Low', 'min'), Close=('Close', 'last'),
        Date_last=('Date', 'max')
    ).reset_index().sort_values('ym')
    
    if len(monthly) >= 1:
        prev_m = monthly.iloc[-1]
        levels = calc_pivot_levels(prev_m['High'], prev_m['Low'], prev_m['Close'])
        result['MN'] = {**levels, 'H': round(prev_m['High'], 2), 'L': round(prev_m['Low'], 2),
                        'C': round(prev_m['Close'], 2), 'date': prev_m['Date_last'].strftime('%Y-%m-%d')}
    
    return result, current_price

def find_confluence_zones(pivots, threshold=20):
    """
    Find pivot levels across different TFs that are within `threshold` points of each other.
    Returns list of clusters, each cluster is a list of {tf, level, price}.
    """
    all_levels = []
    level_names = ['R3', 'R2', 'R1', 'PP', 'S1', 'S2', 'S3']
    for tf, data in pivots.items():
        for lv in level_names:
            if lv in data:
                all_levels.append({'tf': tf, 'level': lv, 'price': data[lv]})
    
    # Cluster levels within threshold
    used = set()
    clusters = []
    for i, a in enumerate(all_levels):
        if i in used:
            continue
        group = [a]
        for j, b in enumerate(all_levels):
            if j <= i or j in used:
                continue
            # Only cluster across DIFFERENT TFs
            if a['tf'] == b['tf']:
                continue
            if abs(a['price'] - b['price']) < threshold:
                group.append(b)
                used.add(j)
        if len(group) >= 2:
            used.add(i)
            clusters.append(group)
    
    # Sort clusters by average price descending
    for cluster in clusters:
        cluster.sort(key=lambda x: x['price'], reverse=True)
    clusters.sort(key=lambda c: sum(x['price'] for x in c) / len(c), reverse=True)
    
    return clusters


def calc_atr(df, period=14):
    """Calculate Average True Range over `period` days."""
    highs = df['High'].values
    lows = df['Low'].values
    closes = df['Close'].values
    n = len(df)
    if n < period + 1:
        return None
    trs = []
    for i in range(n - period, n):
        tr = max(
            highs[i] - lows[i],
            abs(highs[i] - closes[i - 1]),
            abs(lows[i] - closes[i - 1])
        )
        trs.append(tr)
    return sum(trs) / len(trs)


def calc_unified_price_zones(df, pivots, confluences):
    """
    Unified Price Zone — merges Pivot Points + ATR targets + Mean levels
    into a concise set of actionable price zones with confluence scoring.

    Returns list of zone dicts sorted high→low:
      {label, price, sources[], zone_type, confluence_score, description}
    """
    price = df['Close'].values[-1]
    atr14 = calc_atr(df, 14)
    if atr14 is None:
        return [], 0, price

    # MA levels
    ma50 = float(np.mean(df['Close'].values[-50:])) if len(df) >= 50 else None
    ma200 = float(np.mean(df['Close'].values[-200:])) if len(df) >= 200 else None

    # ── Candidate levels ──
    candidates = []

    # ATR-based targets
    candidates.append({'label': 'ATR_BULL', 'price': round(price + atr14, 2),
                        'source': 'ATR +1', 'category': 'bull_target'})
    candidates.append({'label': 'ATR_BULL2', 'price': round(price + 2 * atr14, 2),
                        'source': 'ATR +2', 'category': 'breakout'})
    candidates.append({'label': 'ATR_BEAR', 'price': round(price - atr14, 2),
                        'source': 'ATR -1', 'category': 'bear_target'})
    candidates.append({'label': 'ATR_BEAR2', 'price': round(price - 2 * atr14, 2),
                        'source': 'ATR -2', 'category': 'deep_bear'})

    # Mean reversion targets
    if ma50:
        candidates.append({'label': 'MA50', 'price': round(ma50, 2),
                            'source': 'MA50', 'category': 'mean'})
    if ma200:
        candidates.append({'label': 'MA200', 'price': round(ma200, 2),
                            'source': 'MA200', 'category': 'deep_mean'})

    # Key pivot levels (PP, R1, R2, S1, S2 from all TFs)
    key_levels = ['R2', 'R1', 'PP', 'S1', 'S2']
    for tf in ['D1', 'W1', 'MN']:
        if tf not in pivots:
            continue
        for lv in key_levels:
            if lv in pivots[tf]:
                cat = 'resistance' if lv.startswith('R') else ('pivot' if lv == 'PP' else 'support')
                candidates.append({'label': f'{tf}_{lv}', 'price': pivots[tf][lv],
                                    'source': f'{tf} {lv}', 'category': cat})

    # ── Merge nearby levels (within 0.5 × ATR) into unified zones ──
    merge_threshold = atr14 * 0.5
    candidates.sort(key=lambda x: x['price'], reverse=True)

    zones = []
    used = set()
    for i, c in enumerate(candidates):
        if i in used:
            continue
        group = [c]
        for j, d in enumerate(candidates):
            if j <= i or j in used:
                continue
            if abs(c['price'] - d['price']) < merge_threshold:
                group.append(d)
                used.add(j)
        used.add(i)

        avg_price = round(sum(g['price'] for g in group) / len(group), 2)
        sources = [g['source'] for g in group]
        categories = [g['category'] for g in group]

        # Determine zone type
        if avg_price > price:
            if any('breakout' in c for c in categories):
                zone_type = 'BREAKOUT'
            elif any('bull' in c for c in categories):
                zone_type = 'BULL TARGET'
            else:
                zone_type = 'RESISTANCE'
        elif avg_price < price:
            if any('deep' in c for c in categories):
                zone_type = 'DEEP MEAN'
            elif any('bear' in c for c in categories):
                zone_type = 'BEAR TARGET'
            elif any('support' in c for c in categories):
                zone_type = 'SUPPORT'
            else:
                zone_type = 'PULLBACK'
        else:
            zone_type = 'CURRENT'

        # Confluence score: more sources = stronger zone
        conf_score = len(sources)

        # Check if any confluence cluster overlaps
        for cluster in confluences:
            cluster_avg = sum(x['price'] for x in cluster) / len(cluster)
            if abs(cluster_avg - avg_price) < merge_threshold:
                conf_score += len(cluster)
                break

        # Description
        desc_parts = []
        if conf_score >= 3:
            desc_parts.append('Strong confluence')
        elif conf_score >= 2:
            desc_parts.append('Dual source')
        else:
            desc_parts.append('Single source')
        desc_parts.append(' + '.join(sources))
        dist_pct = round((avg_price - price) / price * 100, 2)
        dist_pts = round(avg_price - price, 2)
        desc_parts.append(f'{dist_pts:+.0f} pts ({dist_pct:+.2f}%)')

        zones.append({
            'label': zone_type,
            'price': avg_price,
            'sources': sources,
            'source_str': ' + '.join(sources),
            'zone_type': zone_type,
            'confluence_score': conf_score,
            'description': ' | '.join(desc_parts),
            'distance_pts': dist_pts,
            'distance_pct': dist_pct,
        })

    # Sort high→low
    zones.sort(key=lambda z: z['price'], reverse=True)

    # Limit: keep max 7 most meaningful zones (top conf_score or closest to price)
    if len(zones) > 7:
        # Always keep zones closest to price and highest confluence
        for z in zones:
            z['_priority'] = z['confluence_score'] * 10 + max(0, 5 - abs(z['distance_pct']))
        zones.sort(key=lambda z: z['_priority'], reverse=True)
        zones = zones[:7]
        for z in zones:
            del z['_priority']
        zones.sort(key=lambda z: z['price'], reverse=True)

    return zones, round(atr14, 2), price


def flatten_zones_for_csv(zones, atr14, current_price):
    """Create flat dict of Unified Price Zone data for CSV columns."""
    flat = {
        'ATR_14d': atr14,
        'UPZ_Count': len(zones),
    }
    # Store up to 7 zones
    for i, z in enumerate(zones):
        flat[f'UPZ_{i+1}_Label'] = z['label']
        flat[f'UPZ_{i+1}_Price'] = z['price']
        flat[f'UPZ_{i+1}_Sources'] = z['source_str']
        flat[f'UPZ_{i+1}_Confluence'] = z['confluence_score']
        flat[f'UPZ_{i+1}_DistPts'] = z['distance_pts']
        flat[f'UPZ_{i+1}_DistPct'] = z['distance_pct']
    # Pad remaining slots
    for i in range(len(zones), 7):
        flat[f'UPZ_{i+1}_Label'] = ''
        flat[f'UPZ_{i+1}_Price'] = ''
        flat[f'UPZ_{i+1}_Sources'] = ''
        flat[f'UPZ_{i+1}_Confluence'] = ''
        flat[f'UPZ_{i+1}_DistPts'] = ''
        flat[f'UPZ_{i+1}_DistPct'] = ''
    return flat

# ── Compute pivots ──
pivots, current_price_pivot = calc_multi_tf_pivots(df)

print(f"\n{'='*55}")
print(f"Multi-TF Pivot Points")
print(f"{'='*55}")
for tf in ['D1', 'W1', 'MN']:
    if tf in pivots:
        p = pivots[tf]
        print(f"  {tf}: PP={p['PP']:.2f}  R1={p['R1']:.2f}  R2={p['R2']:.2f}  R3={p['R3']:.2f}  S1={p['S1']:.2f}  S2={p['S2']:.2f}  S3={p['S3']:.2f}  (H/L/C from {p['date']})")

confluences = find_confluence_zones(pivots, threshold=20)
if confluences:
    print(f"\n⚡ Confluence Zones (within 20 pts):")
    for cluster in confluences:
        avg = sum(x['price'] for x in cluster) / len(cluster)
        tags = " + ".join(f"{x['tf']} {x['level']}({x['price']:.2f})" for x in cluster)
        is_r = any(x['level'].startswith('R') for x in cluster)
        is_s = any(x['level'].startswith('S') for x in cluster)
        zone_type = "Resistance" if is_r and not is_s else ("Support" if is_s and not is_r else "Mixed")
        print(f"  ~{avg:.2f} — {tags} [{zone_type}]")
else:
    print(f"\n  No confluence zones found")

# ── Flatten pivot data for CSV ──
def flatten_pivots_for_csv(pivots, confluences, current_price):
    """Create flat dict of pivot data for CSV columns."""
    flat = {}
    for tf in ['D1', 'W1', 'MN']:
        if tf in pivots:
            for lv in ['PP', 'R1', 'R2', 'R3', 'S1', 'S2', 'S3']:
                flat[f'Pivot_{tf}_{lv}'] = pivots[tf][lv]
            flat[f'Pivot_{tf}_H'] = pivots[tf]['H']
            flat[f'Pivot_{tf}_L'] = pivots[tf]['L']
            flat[f'Pivot_{tf}_C'] = pivots[tf]['C']
            flat[f'Pivot_{tf}_Date'] = pivots[tf]['date']
        else:
            for lv in ['PP', 'R1', 'R2', 'R3', 'S1', 'S2', 'S3']:
                flat[f'Pivot_{tf}_{lv}'] = ''
            flat[f'Pivot_{tf}_H'] = ''
            flat[f'Pivot_{tf}_L'] = ''
            flat[f'Pivot_{tf}_C'] = ''
            flat[f'Pivot_{tf}_Date'] = ''
    
    # Confluence zones: store as pipe-separated string
    # Format: "~avg|TF1 LV1+TF2 LV2|type;~avg|...|type"
    conf_parts = []
    for cluster in confluences:
        avg = sum(x['price'] for x in cluster) / len(cluster)
        tags = "+".join(f"{x['tf']} {x['level']}({x['price']:.2f})" for x in cluster)
        is_r = any(x['level'].startswith('R') for x in cluster)
        is_s = any(x['level'].startswith('S') for x in cluster)
        zone_type = "Resistance" if is_r and not is_s else ("Support" if is_s and not is_r else "Mixed")
        conf_parts.append(f"~{avg:.2f}|{tags}|{zone_type}")
    flat['Confluence_Zones'] = ";".join(conf_parts) if conf_parts else 'None'
    flat['Confluence_Count'] = len(confluences)
    
    # Current price position relative to D1 pivot
    if 'D1' in pivots:
        d1 = pivots['D1']
        if current_price >= d1['R3']:
            flat['Pivot_Position'] = 'Above R3'
        elif current_price >= d1['R2']:
            flat['Pivot_Position'] = 'R2-R3'
        elif current_price >= d1['R1']:
            flat['Pivot_Position'] = 'R1-R2'
        elif current_price >= d1['PP']:
            flat['Pivot_Position'] = 'PP-R1'
        elif current_price >= d1['S1']:
            flat['Pivot_Position'] = 'S1-PP'
        elif current_price >= d1['S2']:
            flat['Pivot_Position'] = 'S2-S1'
        elif current_price >= d1['S3']:
            flat['Pivot_Position'] = 'S3-S2'
        else:
            flat['Pivot_Position'] = 'Below S3'
    else:
        flat['Pivot_Position'] = ''
    
    return flat

pivot_csv = flatten_pivots_for_csv(pivots, confluences, current_price_pivot)

# ── Unified Price Zones (ATR + Pivot + Mean confluence) ──
upz_zones, atr_14d, upz_price = calc_unified_price_zones(df, pivots, confluences)

print(f"\n{'='*55}")
print(f"Unified Price Zones (ATR 14d = {atr_14d:.2f})")
print(f"{'='*55}")
for z in upz_zones:
    conf_stars = '★' * min(z['confluence_score'], 5)
    arrow = '▲' if z['distance_pts'] > 0 else ('▼' if z['distance_pts'] < 0 else '●')
    print(f"  {arrow} {z['label']:<14} {z['price']:>9.2f}  {z['distance_pts']:>+8.0f} pts ({z['distance_pct']:>+.2f}%)  {conf_stars}  [{z['source_str']}]")
print(f"  ● {'NOW':<14} {upz_price:>9.2f}")

upz_csv = flatten_zones_for_csv(upz_zones, atr_14d, upz_price)

# ══════════════════════════════════════════════════════
# Z-SCORE REGIME FILTER
# ══════════════════════════════════════════════════════

def calc_zscore_regime(df, base_idx):
    """
    Calculate Z-Score regime indicators for the given base date index.
    
    Uses Z-Score of closing price relative to rolling mean/std to classify
    the current market regime.
    
    Thresholds are ASYMMETRIC — calibrated for gold's positive drift bias:
      Gold Z-Score 50d historically averages +0.47 (full dataset) to +1.4 (2025-2026).
      Symmetric ±1.5 triggers "Extended" ~45% of the time in trending periods → useless.
    
    Calibrated thresholds (Option B):
      - Extreme Extended: Z >= +2.5  (~6.5% of time)
      - Extended:         Z >= +2.0  (~8.7% additional)
      - Normal:           -1.5 < Z < +2.0  (~77% of time)
      - Depressed:        Z <= -1.5  (~3.3% additional)
      - Extreme Depressed: Z <= -2.0  (~4.2% of time)
    
    Returns dict with Z-Scores at 3 lookback periods (50d, 100d, 200d),
    primary zone classification, and regime description.
    """
    closes = df['Close'].values[:base_idx + 1]
    result = {}
    
    for period, label in [(50, '50d'), (100, '100d'), (200, '200d')]:
        if len(closes) < period:
            result[f'z_{label}'] = None
            continue
        window = closes[-period:]
        mean = np.mean(window)
        std = np.std(window, ddof=1)
        if std == 0:
            result[f'z_{label}'] = 0.0
        else:
            result[f'z_{label}'] = (closes[-1] - mean) / std
    
    # Primary Z-Score = 50d (most responsive, best for regime detection)
    z_primary = result.get('z_50d')
    
    # Asymmetric thresholds — calibrated for gold's upward drift bias
    if z_primary is None:
        result['zone'] = 'N/A'
        result['regime'] = 'Insufficient data for Z-Score'
        result['signal'] = '⚪ N/A'
    elif z_primary >= 2.5:
        result['zone'] = 'Extreme Extended'
        result['regime'] = 'ราคาวิ่งเกิน +2.5σ — pullback risk สูงมาก ควรระวังการเปิด Long ใหม่'
        result['signal'] = '🔴 Extreme Extended (Z≥+2.5)'
    elif z_primary >= 2.0:
        result['zone'] = 'Extended'
        result['regime'] = 'ราคาเหนือ +2.0σ — momentum อาจแรงจริง แต่ pullback risk เพิ่มขึ้น'
        result['signal'] = '🟡 Extended (Z≥+2.0)'
    elif z_primary <= -2.0:
        result['zone'] = 'Extreme Depressed'
        result['regime'] = 'ราคาตกเกิน -2.0σ — oversold สุดโต่ง bounce potential สูง'
        result['signal'] = '🟢 Extreme Depressed (Z≤-2.0)'
    elif z_primary <= -1.5:
        result['zone'] = 'Depressed'
        result['regime'] = 'ราคาต่ำกว่า -1.5σ — oversold zone อาจเป็นจุด mean-revert'
        result['signal'] = '🔵 Depressed (Z≤-1.5)'
    else:
        result['zone'] = 'Normal'
        result['regime'] = 'ราคาอยู่ในกรอบปกติ (-1.5σ ถึง +2.0σ) — momentum score ใช้ได้ตามปกติ'
        result['signal'] = '🟢 Normal'
    
    # Additional context: direction of Z movement (using 50d Z now vs approx 5 days ago)
    if len(closes) >= 55:
        closes_5d_ago = closes[:-5]
        window_5d_ago = closes_5d_ago[-50:]
        mean_ago = np.mean(window_5d_ago)
        std_ago = np.std(window_5d_ago, ddof=1)
        if std_ago > 0:
            z_5d_ago = (closes_5d_ago[-1] - mean_ago) / std_ago
            result['z_delta_5d'] = result['z_50d'] - z_5d_ago
        else:
            result['z_delta_5d'] = 0.0
    else:
        result['z_delta_5d'] = None
    
    return result

# Compute Z-Score for BD2 (latest)
zscore = calc_zscore_regime(df, BD2_idx)

print(f"\n{'='*55}")
print(f"Z-Score Regime Filter")
print(f"{'='*55}")
print(f"  Z-Score 50d:  {zscore['z_50d']:.3f}" if zscore['z_50d'] is not None else "  Z-Score 50d:  N/A")
print(f"  Z-Score 100d: {zscore['z_100d']:.3f}" if zscore['z_100d'] is not None else "  Z-Score 100d: N/A")
print(f"  Z-Score 200d: {zscore['z_200d']:.3f}" if zscore['z_200d'] is not None else "  Z-Score 200d: N/A")
print(f"  Zone:         {zscore['zone']}")
print(f"  Signal:       {zscore['signal']}")
if zscore.get('z_delta_5d') is not None:
    d5 = zscore['z_delta_5d']
    print(f"  Z Delta 5d:   {d5:+.3f} ({'Z rising — extending' if d5 > 0 else 'Z falling — reverting' if d5 < 0 else 'flat'})")

# ══════════════════════════════════════════════════════
# CSV OUTPUT
# ══════════════════════════════════════════════════════

csv_row = {
    'Rank': 1,
    'Ticker': 'GOLD',
    'Net_Score_Avg': round(net_avg, 2),
    'Gross_Score_Avg': round(gross_avg, 2),
    'Net_Score_BD1': round(s1['net'], 2),
    'Net_Score_BD2': round(s2['net'], 2),
    'Score_Delta': round(delta, 2),
    'Tier': momentum_tier,
    'D1_ReturnRank': round(s2['d1'], 2),
    'D2_VolumeRank': round(s2['d2'], 2),
    'D3_RSI': round(s2['d3'], 2),
    'D4_MA': round(s2['d4'], 2),
    'D5_Volatility': round(s2['d5'], 2),
    'D6_External': round(s2['d6'], 2),
    'D6_Raw': s2['d6_raw'],
    'Penalty_Scaled': round(s2['penalty_scaled'], 2),
    'WP_Return_Pct': round(s2['wp_ret'], 2),
    'WP_Volume_Pct': round(s2['wp_vol'], 2),
    'Ret_1Y_Pct': round(s2['penalties']['ret_1y'], 2),
    'Ret_6M_Pct': round(s2['penalties']['ret_6m'], 2),
    'Ret_3M_Pct': round(s2['ret_pctls']['3M']['return'], 2),
    'Ret_1M_Pct': round(s2['ret_pctls']['1M']['return'], 2),
    'Ret_1W_Pct': round(s2['ret_pctls']['1W']['return'], 2),
    'RSI_Value': round(s2['rsi'], 2),
    'MA50': round(s2['ma50'], 2) if s2['ma50'] else '',
    'MA200': round(s2['ma200'], 2) if s2['ma200'] else '',
    'Price': round(s2['price'], 2),
    'Golden_Cross': str(s2['golden_cross']),
    'Volatility_Pct': round(s2['volatility'], 2),
    'Vol_Ratio': round(s2['vol_ratio'], 3),
    'Penalty_Total': s2['penalties']['total'],
    'Penalty_Reversal': s2['penalties']['reversal'],
    'Penalty_DeathCross': s2['penalties']['death_cross'],
    'Warning_Flags': s2['penalties']['flags'] if s2['penalties']['flags'] else 'None',
    'DXY_1M_Pct': round(s2['external']['dxy_1m'], 2) if s2['external']['dxy_1m'] is not None else '',
    'VIX_Level': round(s2['external']['vix_level'], 2) if s2['external']['vix_level'] is not None else '',
    'DXY_Signal': s2['external']['dxy_signal'],
    'VIX_Signal': s2['external']['vix_signal'],
    'News_Top20': 'FALSE',
    'Base_Date_1': s1['date'].strftime('%Y-%m-%d'),
    'Base_Date_2': s2['date'].strftime('%Y-%m-%d'),
    'As_Of_Running': AS_OF,
    # Multi-TF Pivot Points
    **pivot_csv,
    **upz_csv,
    # Z-Score Regime Filter
    'Z_Score_50d': round(zscore['z_50d'], 3) if zscore['z_50d'] is not None else '',
    'Z_Score_100d': round(zscore['z_100d'], 3) if zscore['z_100d'] is not None else '',
    'Z_Score_200d': round(zscore['z_200d'], 3) if zscore['z_200d'] is not None else '',
    'Z_Zone': zscore['zone'],
    'Z_Signal': zscore['signal'],
    'Z_Regime': zscore['regime'],
    'Z_Delta_5d': round(zscore['z_delta_5d'], 3) if zscore.get('z_delta_5d') is not None else '',
}

csv_df = pd.DataFrame([csv_row])
csv_fixed = os.path.join(base_dir, 'output_momentum_gold.csv')
csv_ts = os.path.join(base_dir, f'output_momentum_gold_{TS_FILE}.csv')
csv_df.to_csv(csv_fixed, index=False, encoding='utf-8')
csv_df.to_csv(csv_ts, index=False, encoding='utf-8')
print(f"\nCSV saved: {csv_fixed}")
print(f"CSV saved: {csv_ts}")

# ══════════════════════════════════════════════════════
# SCORE HISTORY — append daily (for exhaustion/scenario analysis)
# ══════════════════════════════════════════════════════

history_row = {
    'Date': s2['date'].strftime('%Y-%m-%d'),
    'Price': round(s2['price'], 2),
    'Net_Score': round(s2['net'], 2),
    'Gross_Score': round(s2['gross'], 2),
    'D1_Return': round(s2['d1'], 2),
    'D2_Volume': round(s2['d2'], 2),
    'D3_RSI': round(s2['d3'], 2),
    'D4_MA': round(s2['d4'], 2),
    'D5_DirVol': round(s2['d5'], 2),
    'D6_External': round(s2['d6'], 2),
    'Vol_Ratio': round(s2['vol_ratio'], 3),
    'RSI': round(s2['rsi'], 2),
    'Volatility_Pct': round(s2['volatility'], 2),
    'Penalty_Scaled': round(s2['penalty_scaled'], 2),
    'Ret_1W': round(s2['ret_pctls']['1W']['return'], 2),
    'Ret_1M': round(s2['ret_pctls']['1M']['return'], 2),
    'Ret_3M': round(s2['ret_pctls']['3M']['return'], 2),
    'Golden_Cross': str(s2['golden_cross']),
    'Z_Score_50d': round(zscore['z_50d'], 3) if zscore['z_50d'] is not None else '',
    'Z_Zone': zscore['zone'],
    'Z_Delta_5d': round(zscore['z_delta_5d'], 3) if zscore.get('z_delta_5d') is not None else '',
    'Warning_Flags': s2['penalties']['flags'] if s2['penalties']['flags'] else 'None',
    'Tier': momentum_tier,
    'As_Of_Running': AS_OF,
}

history_path = os.path.join(base_dir, 'score_history.csv')
history_df = pd.DataFrame([history_row])

if os.path.exists(history_path):
    existing = pd.read_csv(history_path, encoding='utf-8')
    # Ensure string columns don't become float64 from empty values
    for col in ['Exhaust_Scenario', 'Warning_Flags', 'Tier', 'Z_Zone', 'Golden_Cross', 'As_Of_Running']:
        if col in existing.columns:
            existing[col] = existing[col].fillna('').astype(str)
    # ไม่ซ้ำวันเดียวกัน — ถ้ารันซ้ำวันเดิม overwrite แถวนั้น
    existing = existing[existing['Date'] != history_row['Date']]
    history_df = pd.concat([existing, history_df], ignore_index=True)
    history_df = history_df.sort_values('Date').reset_index(drop=True)

history_df.to_csv(history_path, index=False, encoding='utf-8')
print(f"Score history: {history_path} ({len(history_df)} rows)")

# ══════════════════════════════════════════════════════
# EXHAUSTION DETECTION (from score_history backtest)
# ══════════════════════════════════════════════════════

exhaust_result = {
    'scenario': 'None',
    'label': '',
    'action_override': '',
    'net_5d_change': '',
    'max_10d': '',
    'min_10d': '',
    'd5_shift_5d': '',
}

if len(history_df) >= 6:
    h = history_df.copy()
    h['Net_Score'] = pd.to_numeric(h['Net_Score'], errors='coerce')
    h['D5_DirVol'] = pd.to_numeric(h['D5_DirVol'], errors='coerce')
    h['Z_Score_50d'] = pd.to_numeric(h['Z_Score_50d'], errors='coerce')

    current = h.iloc[-1]
    net_now = current['Net_Score']
    d5_now = current['D5_DirVol']
    z_now = current['Z_Score_50d'] if pd.notna(current['Z_Score_50d']) else 0

    # Net score 5d ago
    net_5d_ago = h.iloc[-6]['Net_Score'] if len(h) >= 6 else net_now
    net_5d_change = net_now - net_5d_ago

    # Max/Min Net in last 10 days
    last10 = h['Net_Score'].tail(min(10, len(h)))
    max_10d = last10.max()
    min_10d = last10.min()

    # D5 shift in 5 days
    d5_5d_ago = h.iloc[-6]['D5_DirVol'] if len(h) >= 6 else d5_now
    d5_shift = d5_now - d5_5d_ago

    exhaust_result['net_5d_change'] = round(net_5d_change, 2)
    exhaust_result['max_10d'] = round(max_10d, 2)
    exhaust_result['min_10d'] = round(min_10d, 2)
    exhaust_result['d5_shift_5d'] = round(d5_shift, 2)

    # ── sc13: Bull Exhaustion ──
    # Z extended + score still high + momentum fading
    # Backtest: Fwd5d = -0.29% (ONLY pattern worse than baseline)
    sc13 = z_now >= 2.0 and net_now >= 70 and net_5d_change < 0

    # ── sc14: Topping Alert ──
    # Was very strong recently but dropped hard
    # Backtest: Fwd5d = +1.83% (bounce likely, don't panic)
    sc14 = max_10d >= 80 and net_5d_change < -8 and not sc13

    # ── sc15: Bear Exhaustion / Bottoming ──
    # Was low recently + starting to recover
    # Backtest: Fwd10d = +4-7% (contrarian buy)
    sc15 = min_10d < 50 and net_5d_change > 3

    # ── sc16: D5 Volatility Regime Shift ──
    # D5 changed ≥50pts in 5 days → market structure change
    # Backtest: Fwd10d = +4.06%
    sc16 = abs(d5_shift) >= 50 and not sc13 and not sc14 and not sc15

    if sc13:
        exhaust_result['scenario'] = 'Bull Exhaustion'
        exhaust_result['label'] = '🔥 Bull Exhaustion: Z extended + momentum fading → HOLD'
        exhaust_result['action_override'] = 'HOLD'
    elif sc15:
        exhaust_result['scenario'] = 'Bear Exhaustion'
        exhaust_result['label'] = '🔋 Bear Exhaustion: Selling exhausted, bounce likely → BUY'
        exhaust_result['action_override'] = 'BUY'
    elif sc14:
        exhaust_result['scenario'] = 'Topping'
        exhaust_result['label'] = '🏔️ Topping: Score collapsed from recent high → HOLD (bounce likely)'
        exhaust_result['action_override'] = 'HOLD'
    elif sc16:
        exhaust_result['scenario'] = 'Vol Shift'
        exhaust_result['label'] = '⚡ Vol Regime Shift: D5 changed ' + str(round(d5_shift)) + 'pts → HOLD'
        exhaust_result['action_override'] = 'HOLD'

    print(f"\n── Exhaustion Detection ──")
    print(f"  Net 5d Δ:    {net_5d_change:+.2f}")
    print(f"  Max 10d:     {max_10d:.2f}  |  Min 10d: {min_10d:.2f}")
    print(f"  D5 shift 5d: {d5_shift:+.0f}  (was {d5_5d_ago:.0f} → now {d5_now:.0f})")
    print(f"  Z-Score:     {z_now:.3f}")
    if exhaust_result['scenario'] != 'None':
        print(f"  >>> {exhaust_result['label']}")
    else:
        print(f"  >>> No exhaustion signal")

# Add exhaustion columns to main CSV
csv_row['Exhaust_Scenario'] = exhaust_result['scenario']
csv_row['Exhaust_Action'] = exhaust_result['action_override']
csv_row['Net_5d_Change'] = exhaust_result['net_5d_change']
csv_row['Max_10d'] = exhaust_result['max_10d']
csv_row['Min_10d'] = exhaust_result['min_10d']
csv_row['D5_Shift_5d'] = exhaust_result['d5_shift_5d']

# Re-save CSVs with exhaustion columns
csv_df = pd.DataFrame([csv_row])
csv_df.to_csv(csv_fixed, index=False, encoding='utf-8')
csv_df.to_csv(csv_ts, index=False, encoding='utf-8')
print(f"\nCSV updated with exhaustion: {csv_fixed}")

# Also update the history row with exhaustion info
history_df.loc[history_df['Date'] == history_row['Date'], 'Exhaust_Scenario'] = exhaust_result['scenario']
history_df.to_csv(history_path, index=False, encoding='utf-8')

# ══════════════════════════════════════════════════════
# EXCEL OUTPUT
# ══════════════════════════════════════════════════════

wb = Workbook()

# ── STYLES ──
gold_fill = PatternFill('solid', fgColor='FFD700')
green_fill = PatternFill('solid', fgColor='C6EFCE')
red_fill = PatternFill('solid', fgColor='FFC7CE')
blue_fill = PatternFill('solid', fgColor='BDD7EE')
purple_fill = PatternFill('solid', fgColor='E2D0F8')
gray_fill = PatternFill('solid', fgColor='D9D9D9')
dark_fill = PatternFill('solid', fgColor='333333')
header_font = Font(bold=True, size=11, color='FFFFFF')
title_font = Font(bold=True, size=14, color='333333')
big_font = Font(bold=True, size=18, color='1F4E79')
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header_row(ws, row, cols, fill=None):
    f = fill or PatternFill('solid', fgColor='1F4E79')
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = f
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

def style_cell(ws, row, col, fmt=None):
    cell = ws.cell(row=row, column=col)
    cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center')
    if fmt: cell.number_format = fmt
    return cell

# ════════════ SHEET 1: SUMMARY ════════════
ws1 = wb.active
ws1.title = "Gold Momentum v3.0"
ws1.sheet_properties.tabColor = "FFD700"

ws1.merge_cells('A1:H1')
ws1['A1'] = "🥇 Gold Momentum Score v3.0 — 100-Scale Edition"
ws1['A1'].font = big_font
ws1['A1'].alignment = Alignment(horizontal='center')

ws1.merge_cells('A2:H2')
ws1['A2'] = f"Run: {AS_OF}  |  BD1: {s1['date'].strftime('%Y-%m-%d')}  |  BD2: {s2['date'].strftime('%Y-%m-%d')}"
ws1['A2'].font = Font(size=10, color='666666')
ws1['A2'].alignment = Alignment(horizontal='center')

ws1.merge_cells('A4:B4')
ws1['A4'] = "Net Score Avg"
ws1['A4'].font = Font(bold=True, size=12)
ws1['C4'] = round(net_avg, 2)
ws1['C4'].font = Font(bold=True, size=16, color='1F4E79')
ws1['C4'].number_format = '0.00'
ws1['D4'] = momentum_tier
ws1['D4'].font = Font(bold=True, size=14)
ws1['D4'].fill = gold_fill
ws1['F4'] = f"Price: ${s2['price']:,.1f}"
ws1['F4'].font = Font(bold=True, size=12)

# Summary table
row = 6
headers = ['Metric', 'BD1', 'BD2', 'Average', 'Delta']
for i, h in enumerate(headers):
    ws1.cell(row=row, column=i+1, value=h)
style_header_row(ws1, row, len(headers))

data_rows = [
    ['Net Score', s1['net'], s2['net'], net_avg, delta],
    ['Gross Score (Avg D1-D6)', s1['gross'], s2['gross'], gross_avg, s2['gross']-s1['gross']],
    ['D1 Return Rank (/100)', s1['d1'], s2['d1'], (s1['d1']+s2['d1'])/2, s2['d1']-s1['d1']],
    ['D2 Volume Rank (/100)', s1['d2'], s2['d2'], (s1['d2']+s2['d2'])/2, s2['d2']-s1['d2']],
    ['D3 RSI (/100)', s1['d3'], s2['d3'], (s1['d3']+s2['d3'])/2, s2['d3']-s1['d3']],
    ['D4 MA Trend (/100)', s1['d4'], s2['d4'], (s1['d4']+s2['d4'])/2, s2['d4']-s1['d4']],
    ['D5 Dir.Volatility (/100)', s1['d5'], s2['d5'], (s1['d5']+s2['d5'])/2, s2['d5']-s1['d5']],
    ['D6 External (/100)', s1['d6'], s2['d6'], (s1['d6']+s2['d6'])/2, s2['d6']-s1['d6']],
    ['Penalty (raw)', s1['penalties']['total'], s2['penalties']['total'], (s1['penalties']['total']+s2['penalties']['total'])/2, s2['penalties']['total']-s1['penalties']['total']],
    ['Penalty (scaled /100)', s1['penalty_scaled'], s2['penalty_scaled'], (s1['penalty_scaled']+s2['penalty_scaled'])/2, s2['penalty_scaled']-s1['penalty_scaled']],
]
for r, row_data in enumerate(data_rows):
    rn = row + 1 + r
    for c, val in enumerate(row_data):
        cell = style_cell(ws1, rn, c+1, '0.00' if isinstance(val, float) else None)
        cell.value = val
        if c == 4 and isinstance(val, (int, float)):
            cell.font = Font(color='006100' if val >= 0 else '9C0006')

# External Context detail
ext_row = row + 1 + len(data_rows) + 2
ws1.cell(row=ext_row, column=1, value="📊 External Context Detail (BD2)").font = Font(bold=True, size=12)
ext_row += 1
ext_headers = ['Factor', 'Score', 'Signal', 'Value']
for i, h in enumerate(ext_headers):
    ws1.cell(row=ext_row, column=i+1, value=h)
style_header_row(ws1, ext_row, len(ext_headers), PatternFill('solid', fgColor='7030A0'))

ext_row += 1
dxy_val = f"{s2['external']['dxy_1m']:.2f}%" if s2['external']['dxy_1m'] is not None else "N/A"
for c, v in enumerate(['DXY Divergence', s2['external']['dxy_score'], s2['external']['dxy_signal'], f"DXY 1M: {dxy_val}"]):
    style_cell(ws1, ext_row, c+1).value = v

ext_row += 1
vix_val = f"{s2['external']['vix_level']:.2f}" if s2['external']['vix_level'] is not None else "N/A"
for c, v in enumerate(['VIX Regime', s2['external']['vix_score'], s2['external']['vix_signal'], f"VIX: {vix_val}"]):
    style_cell(ws1, ext_row, c+1).value = v

ext_row += 1
for c, v in enumerate(['D6 Total', s2['d6'], '', '']):
    cell = style_cell(ws1, ext_row, c+1)
    cell.value = v
    if c == 1:
        cell.font = Font(bold=True, size=12, color='006100' if s2['d6'] >= 0 else '9C0006')

# Return performance
ret_row = ext_row + 2
ws1.cell(row=ret_row, column=1, value="Return Performance (BD2)").font = Font(bold=True, size=12)
ret_row += 1
ret_headers = ['Period', 'Days', 'Return %', 'Rolling Percentile', 'Weight']
for i, h in enumerate(ret_headers):
    ws1.cell(row=ret_row, column=i+1, value=h)
style_header_row(ws1, ret_row, len(ret_headers))

for p in WEIGHT_ORDER:
    ret_row += 1
    vals = [p, LOOKBACK[p], s2['ret_pctls'][p]['return'], s2['ret_pctls'][p]['percentile'], WEIGHTS[p]*100]
    for c, v in enumerate(vals):
        cell = style_cell(ws1, ret_row, c+1, '0.00')
        cell.value = round(v, 2) if isinstance(v, float) else v

for c in range(1, 9):
    ws1.column_dimensions[get_column_letter(c)].width = 22

# ════════════ SHEET 2: RETURN PERCENTILE DETAIL ════════════
ws2 = wb.create_sheet("Return Percentile Detail")
ws2.sheet_properties.tabColor = "4472C4"

ws2['A1'] = "📊 Return Percentile Detail (Rolling 252-day)"
ws2['A1'].font = title_font
ws2.merge_cells('A1:F1')

for bd_label, sc, start_row in [("Base Date 1", s1, 3), ("Base Date 2", s2, 10)]:
    ws2.cell(row=start_row, column=1, value=f"{bd_label}: {sc['date'].strftime('%Y-%m-%d')}").font = Font(bold=True, size=11)
    hr = start_row + 1
    cols = ['Period', 'Days', 'Weight', 'Raw Return %', 'Percentile Rank', 'Weighted Contribution']
    for i, h in enumerate(cols):
        ws2.cell(row=hr, column=i+1, value=h)
    style_header_row(ws2, hr, len(cols))
    
    for p in WEIGHT_ORDER:
        hr += 1
        pdata = sc['ret_pctls'][p]
        contrib = pdata['percentile'] * WEIGHTS[p]
        vals = [p, LOOKBACK[p], f"{WEIGHTS[p]*100:.0f}%", round(pdata['return'], 2), round(pdata['percentile'], 2), round(contrib, 2)]
        for c, v in enumerate(vals):
            style_cell(ws2, hr, c+1).value = v
    
    hr += 1
    ws2.cell(row=hr, column=1, value="Total WP_Return").font = Font(bold=True)
    ws2.cell(row=hr, column=5, value=round(sc['wp_ret'], 2)).font = Font(bold=True)
    hr += 1
    ws2.cell(row=hr, column=1, value="D1 Score (WP/100 × 20)").font = Font(bold=True)
    ws2.cell(row=hr, column=5, value=round(sc['d1'], 2)).font = Font(bold=True, color='1F4E79')

for c in range(1, 7):
    ws2.column_dimensions[get_column_letter(c)].width = 22

# ════════════ SHEET 3: VOLUME PERCENTILE DETAIL ════════════
ws3 = wb.create_sheet("Volume Percentile Detail")
ws3.sheet_properties.tabColor = "70AD47"

ws3['A1'] = "📊 Volume Percentile Detail (Rolling 252-day)"
ws3['A1'].font = title_font
ws3.merge_cells('A1:F1')

for bd_label, sc, start_row in [("Base Date 1", s1, 3), ("Base Date 2", s2, 10)]:
    ws3.cell(row=start_row, column=1, value=f"{bd_label}: {sc['date'].strftime('%Y-%m-%d')}").font = Font(bold=True)
    hr = start_row + 1
    cols = ['Period', 'Days', 'Weight', 'Cum Volume', 'Percentile Rank', 'Weighted Contribution']
    for i, h in enumerate(cols):
        ws3.cell(row=hr, column=i+1, value=h)
    style_header_row(ws3, hr, len(cols))
    
    for p in WEIGHT_ORDER:
        hr += 1
        pdata = sc['vol_pctls'][p]
        contrib = pdata['percentile'] * WEIGHTS[p]
        vals = [p, LOOKBACK[p], f"{WEIGHTS[p]*100:.0f}%", round(pdata['volume'], 0), round(pdata['percentile'], 2), round(contrib, 2)]
        for c, v in enumerate(vals):
            style_cell(ws3, hr, c+1).value = v
    
    hr += 1
    ws3.cell(row=hr, column=1, value="Total WP_Volume").font = Font(bold=True)
    ws3.cell(row=hr, column=5, value=round(sc['wp_vol'], 2)).font = Font(bold=True)
    hr += 1
    ws3.cell(row=hr, column=1, value="D2 Score (WP/100 × 20)").font = Font(bold=True)
    ws3.cell(row=hr, column=5, value=round(sc['d2'], 2)).font = Font(bold=True, color='1F4E79')

for c in range(1, 7):
    ws3.column_dimensions[get_column_letter(c)].width = 22

# ════════════ SHEET 4: WARNING FLAGS + EXTERNAL ════════════
ws4 = wb.create_sheet("🚨 Warnings & External")
ws4.sheet_properties.tabColor = "FF0000"

ws4['A1'] = "🚨 Warning Flags, Penalties & External Context"
ws4['A1'].font = title_font
ws4.merge_cells('A1:F1')

for bd_label, sc, start_row in [("Base Date 2 (Latest)", s2, 3), ("Base Date 1", s1, 14)]:
    ws4.cell(row=start_row, column=1, value=f"{bd_label}: {sc['date'].strftime('%Y-%m-%d')}").font = Font(bold=True, size=11)
    hr = start_row + 1
    cols = ['Penalty Type', 'Pts', 'Condition', 'Current Values', 'Status']
    for i, h in enumerate(cols):
        ws4.cell(row=hr, column=i+1, value=h)
    style_header_row(ws4, hr, len(cols), PatternFill('solid', fgColor='C00000'))
    
    pen = sc['penalties']
    hr += 1
    rev_status = "TRIGGERED" if pen['reversal'] != 0 else "CLEAN"
    vals = ['Momentum Reversal', pen['reversal'],
            '1Y>0 & 1M<0 & 1W<0 (mild) / 1Y>20 & 1M<-5 & 1W<-3 (strong)',
            f"1Y={pen['ret_1y']:.1f}% 6M={pen['ret_6m']:.1f}% 1M={pen['ret_1m']:.1f}% 1W={pen['ret_1w']:.1f}%",
            rev_status]
    for c, v in enumerate(vals):
        cell = style_cell(ws4, hr, c+1)
        cell.value = v
        if c == 4:
            cell.fill = red_fill if rev_status == "TRIGGERED" else green_fill
    
    hr += 1
    dc_status = "TRIGGERED" if pen['death_cross'] != 0 else "CLEAN"
    ma50_v = sc['ma50'] if sc['ma50'] else 0
    ma200_v = sc['ma200'] if sc['ma200'] else 0
    vals = ['Death Cross', pen['death_cross'], 'MA50 < MA200',
            f"MA50={ma50_v:.1f} MA200={ma200_v:.1f} Price={sc['price']:.1f}", dc_status]
    for c, v in enumerate(vals):
        cell = style_cell(ws4, hr, c+1)
        cell.value = v
        if c == 4:
            cell.fill = red_fill if dc_status == "TRIGGERED" else green_fill

    # External Context in same sheet
    hr += 2
    ws4.cell(row=hr, column=1, value="External Context (D6)").font = Font(bold=True, size=11, color='7030A0')
    hr += 1
    ext_cols = ['Factor', 'Pts', 'Signal', 'Value', 'Status']
    for i, h in enumerate(ext_cols):
        ws4.cell(row=hr, column=i+1, value=h)
    style_header_row(ws4, hr, len(ext_cols), PatternFill('solid', fgColor='7030A0'))
    
    ext = sc['external']
    hr += 1
    dxy_val = f"{ext['dxy_1m']:.2f}%" if ext['dxy_1m'] is not None else "N/A"
    dxy_status = "BONUS" if ext['dxy_score'] > 0 else ("PENALTY" if ext['dxy_score'] < 0 else "NEUTRAL")
    for c, v in enumerate(['DXY Divergence', ext['dxy_score'], ext['dxy_signal'], f"DXY 1M: {dxy_val}", dxy_status]):
        cell = style_cell(ws4, hr, c+1)
        cell.value = v
        if c == 4:
            cell.fill = green_fill if ext['dxy_score'] > 0 else (red_fill if ext['dxy_score'] < 0 else gray_fill)

    hr += 1
    vix_val = f"{ext['vix_level']:.2f}" if ext['vix_level'] is not None else "N/A"
    vix_status = "BONUS" if ext['vix_score'] > 0 else ("PENALTY" if ext['vix_score'] < 0 else "NEUTRAL")
    for c, v in enumerate(['VIX Regime', ext['vix_score'], ext['vix_signal'], f"VIX: {vix_val}", vix_status]):
        cell = style_cell(ws4, hr, c+1)
        cell.value = v
        if c == 4:
            cell.fill = green_fill if ext['vix_score'] > 0 else (red_fill if ext['vix_score'] < 0 else gray_fill)

    hr += 1
    ws4.cell(row=hr, column=1, value=f"Total Penalty: {pen['total']}  |  D6: {ext['d6_total']:+d}  |  Net Score: {sc['net']:.2f}").font = Font(bold=True)

for c in range(1, 6):
    ws4.column_dimensions[get_column_letter(c)].width = 35

# ════════════ SHEET 5: METHODOLOGY v2.0 ════════════
ws5 = wb.create_sheet("Methodology v3.0")
ws5.sheet_properties.tabColor = "7030A0"

ws5['A1'] = "📘 Methodology — Gold Momentum Scoring v3.0 (100-Scale)"
ws5['A1'].font = title_font
ws5.merge_cells('A1:C1')

methods = [
    ("D1: Return Rank (0-100)", "Rolling Percentile of returns vs self over 252 days\nWeights: 1Y=30%, 6M=25%, 3M=20%, 1M=15%, 1W=10%\nScore = Weighted Percentile (already 0-100)"),
    ("D2: Volume Rank (0-100)", "Rolling Percentile of cumulative volume vs self over 252 days\nSame weights as D1\nScore = Weighted Percentile (already 0-100)"),
    ("D3: RSI (0-100)", "14-day RSI\n50-70→100, 40-49→80, 71-80→70\n30-39→60, >80→50, <30→30"),
    ("D4: MA Trend (0-100)", "Price>MA50→+35, Price>MA200→+35, Golden Cross(MA50>MA200)→+30\nMax: 100 pts"),
    ("D5: Directional Volatility (0-100)", "21-day directional volatility\nvol_ratio = downside_vol / upside_vol\n≤0.6→100, ≤0.8→85, ≤1.0→70, ≤1.2→55\n≤1.5→40, ≤2.0→20, >2.0→10"),
    ("D6: External Context (0-100)", "Raw score ±10 mapped to 0-100 (center=50)\n-10→0, 0→50, +10→100\nPart A — DXY Divergence (±5)\nPart B — VIX Regime (±5)"),
    ("Penalty System (0 to -14)", "Same triggers as before, scaled from /110 to /100:\nPenalty_Scaled = Penalty_Raw × (100/110)\nMild Reversal: ~-4.5 | Strong Reversal: ~-9.1 | Death Cross: ~-4.5\nMax: ~-13.6 (capped)"),
    ("Net Score Formula", "Gross = Average(D1, D2, D3, D4, D5, D6)  [all /100]\nNet = Gross + Penalty_Scaled\nAll dimensions weighted equally (1/6 each)"),
    ("Data Sources", f"Gold: Yahoo Finance GC=F | DXY: Yahoo Finance DX-Y.NYB\nVIX: Yahoo Finance ^VIX\nRepo: github.com/jptrustlearning/gold\nRange: {df.iloc[0]['Date'].strftime('%Y-%m-%d')} to {df.iloc[-1]['Date'].strftime('%Y-%m-%d')}")
]

row = 3
for title, desc in methods:
    ws5.cell(row=row, column=1, value=title).font = Font(bold=True, size=11, color='7030A0')
    ws5.cell(row=row, column=2, value=desc).alignment = Alignment(wrap_text=True)
    row += 3

ws5.column_dimensions['A'].width = 35
ws5.column_dimensions['B'].width = 75

# ── SAVE EXCEL ──
excel_path = os.path.join(base_dir, 'Gold_Momentum_v3.0.xlsx')
wb.save(excel_path)
print(f"\nExcel saved: {excel_path}")

print(f"\n✅ All outputs generated successfully!")
print(f"   CSV: output_momentum_gold.csv + output_momentum_gold_{TS_FILE}.csv")
print(f"   Excel: Gold_Momentum_v3.0.xlsx")
